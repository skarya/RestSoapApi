"""
reporter.py
Generates a richly styled Excel report using openpyxl.

Sheet 1 — "Summary Dashboard"  : KPI cards, pass rate bar, run metadata
Sheet 2 — "Test Results"       : Full per-row details with conditional formatting
"""
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Colour palette ────────────────────────────────────────────────────────────
C_NAVY        = "1F2D3D"
C_WHITE       = "FFFFFF"
C_PASS_DARK   = "1A6B3C"
C_PASS_LIGHT  = "D6F5E3"
C_FAIL_DARK   = "9B1C1C"
C_FAIL_LIGHT  = "FDE8E8"
C_ERROR_DARK  = "7B3F00"
C_ERROR_LIGHT = "FFF3CD"
C_REST        = "1565C0"
C_SOAP        = "6A1B9A"
C_GET         = "00796B"
C_POST        = "E65100"
C_PUT         = "1976D2"
C_DELETE      = "C62828"
C_PATCH       = "6D4C41"
C_ROW_ALT     = "F2F6FC"
C_HEADER_BG   = "1F2D3D"
C_STATUS_2XX  = "C8E6C9"
C_STATUS_4XX  = "FFE0B2"
C_STATUS_5XX  = "FFCDD2"
C_TIME_OK     = "C8E6C9"
C_TIME_WARN   = "FFE0B2"
C_TIME_SLOW   = "FFCDD2"
C_GREY_BG     = "F5F5F5"
C_PAYLOAD_BG  = "FFFDE7"
C_RESP_BG     = "E3F2FD"
C_BORDER      = "CCCCCC"
C_TITLE_GOLD  = "F9A825"


def _thin_border():
    side = Side(style="thin", color=C_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color=C_NAVY, size=10, italic=False, name="Calibri"):
    return Font(name=name, bold=bold, color=color, size=size, italic=italic)


def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _method_color(method: str) -> str:
    m = method.upper()
    return {"GET": C_GET, "POST": C_POST, "PUT": C_PUT,
            "DELETE": C_DELETE, "PATCH": C_PATCH}.get(m, C_NAVY)


def _status_color(code: int) -> str:
    if 200 <= code < 300:
        return C_STATUS_2XX
    elif 400 <= code < 500:
        return C_STATUS_4XX
    elif code >= 500:
        return C_STATUS_5XX
    return C_GREY_BG


def _time_color(ms: float) -> str:
    if ms <= 500:
        return C_TIME_OK
    elif ms <= 2000:
        return C_TIME_WARN
    return C_TIME_SLOW


def _set_col_width(ws, col_idx: int, width: float):
    ws.column_dimensions[get_column_letter(col_idx)].width = width


def _set_row_height(ws, row_idx: int, height: float):
    ws.row_dimensions[row_idx].height = height


# ── Sheet 1: Summary Dashboard ────────────────────────────────────────────────

def _write_summary_sheet(ws, results: list, input_file: str, mode: str,
                          start_time: datetime, end_time: datetime):
    ws.sheet_view.showGridLines = False

    total   = len(results)
    passed  = sum(1 for r in results if r.get("pass_fail") == "PASS")
    failed  = sum(1 for r in results if r.get("pass_fail") == "FAIL")
    errors  = sum(1 for r in results if r.get("pass_fail") == "ERROR")
    rest_c  = sum(1 for r in results if r.get("api_type")  == "REST")
    soap_c  = sum(1 for r in results if r.get("api_type")  == "SOAP")
    pass_pct = round((passed / total * 100), 1) if total else 0
    duration = (end_time - start_time).total_seconds()

    # ── Title banner ──
    ws.merge_cells("B2:J3")
    cell = ws["B2"]
    cell.value     = "🚀  API TEST EXECUTION REPORT"
    cell.font      = Font(name="Calibri", bold=True, size=22, color=C_TITLE_GOLD)
    cell.fill      = _fill(C_NAVY)
    cell.alignment = _align("center", "center")
    _set_row_height(ws, 2, 45)
    _set_row_height(ws, 3, 45)

    # ── Subtitle / metadata row ──
    ws.merge_cells("B4:J4")
    meta = ws["B4"]
    meta.value = (
        f"Input: {os.path.basename(input_file)}   |   "
        f"Mode: {mode}   |   "
        f"Run: {start_time.strftime('%Y-%m-%d %H:%M:%S')}   |   "
        f"Duration: {duration:.1f}s"
    )
    meta.font      = _font(size=10, color="AAAAAA", italic=True)
    meta.fill      = _fill(C_NAVY)
    meta.alignment = _align("center", "center")
    _set_row_height(ws, 4, 22)

    # ── KPI Cards (row 6–9) ──
    _kpi_card(ws, row=6, col=2, label="TOTAL",  value=total,  bg="34495E", fg=C_WHITE)
    _kpi_card(ws, row=6, col=4, label="PASSED", value=passed, bg=C_PASS_DARK, fg=C_WHITE)
    _kpi_card(ws, row=6, col=6, label="FAILED", value=failed, bg=C_FAIL_DARK, fg=C_WHITE)
    _kpi_card(ws, row=6, col=8, label="ERRORS", value=errors, bg=C_ERROR_DARK, fg=C_WHITE)
    _set_row_height(ws, 6, 22)
    _set_row_height(ws, 7, 38)
    _set_row_height(ws, 8, 18)
    _set_row_height(ws, 9, 12)

    # ── Pass rate bar (row 11–12) ──
    ws.merge_cells("B11:C11")
    label = ws["B11"]
    label.value     = "Pass Rate"
    label.font      = _font(bold=True, size=11)
    label.alignment = _align("right", "center")

    ws.merge_cells("D11:D11")
    pct_cell = ws["D11"]
    pct_cell.value     = f"{pass_pct}%"
    pct_cell.font      = _font(bold=True, size=12, color=C_PASS_DARK)
    pct_cell.alignment = _align("center", "center")

    # Draw bar (10 blocks = E11 to N11)
    filled_blocks = round(pass_pct / 10)
    for i, col_letter in enumerate([get_column_letter(c) for c in range(5, 15)]):
        bar_cell = ws[f"{col_letter}11"]
        bar_cell.fill      = _fill(C_PASS_DARK if i < filled_blocks else "E0E0E0")
        bar_cell.border    = _thin_border()
    _set_row_height(ws, 11, 22)

    # ── API type breakdown (row 13) ──
    ws.merge_cells("B13:C13")
    t = ws["B13"]
    t.value = "REST calls"
    t.font      = _font(bold=True, color=C_WHITE)
    t.fill      = _fill(C_REST)
    t.alignment = _align("center", "center")
    t.border    = _thin_border()

    ws["D13"].value     = rest_c
    ws["D13"].font      = _font(bold=True, size=14)
    ws["D13"].alignment = _align("center", "center")

    ws.merge_cells("F13:G13")
    s = ws["F13"]
    s.value     = "SOAP calls"
    s.font      = _font(bold=True, color=C_WHITE)
    s.fill      = _fill(C_SOAP)
    s.alignment = _align("center", "center")
    s.border    = _thin_border()

    ws["H13"].value     = soap_c
    ws["H13"].font      = _font(bold=True, size=14)
    ws["H13"].alignment = _align("center", "center")
    _set_row_height(ws, 13, 28)

    # Column widths for dashboard
    for col, w in [(2,  14), (3, 2), (4, 10), (5, 2), (6, 14),
                   (7, 2),   (8, 10), (9, 2), (10, 14)]:
        _set_col_width(ws, col, w)


def _kpi_card(ws, row: int, col: int, label: str, value: int, bg: str, fg: str):
    ws.merge_cells(
        start_row=row, start_column=col, end_row=row, end_column=col + 1
    )
    lc = ws.cell(row=row, column=col)
    lc.value     = label
    lc.font      = Font(name="Calibri", bold=True, size=10, color=fg)
    lc.fill      = _fill(bg)
    lc.alignment = _align("center", "center")
    lc.border    = _thin_border()

    ws.merge_cells(
        start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1
    )
    vc = ws.cell(row=row + 1, column=col)
    vc.value     = value
    vc.font      = Font(name="Calibri", bold=True, size=28, color=bg)
    vc.fill      = _fill(C_WHITE)
    vc.alignment = _align("center", "center")
    vc.border    = _thin_border()


# ── Sheet 2: Test Results ─────────────────────────────────────────────────────

COLUMNS = [
    ("#",               5,  "center"),
    ("Test Case ID",   20,  "left"),
    ("Type",            8,  "center"),
    ("Method",         10,  "center"),
    ("Endpoint",       42,  "left"),
    ("Payload Sent",   36,  "left"),
    ("Status Code",    13,  "center"),
    ("Response Body",  52,  "left"),
    ("Response (ms)",  15,  "center"),
    ("Status",         12,  "center"),
]


def _write_results_sheet(ws, results: list):
    ws.sheet_view.showGridLines = False

    # ── Header row ──
    _set_row_height(ws, 1, 40)
    for col_idx, (col_name, col_width, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value     = col_name.upper()
        cell.font      = Font(name="Calibri", bold=True, size=11, color=C_WHITE)
        cell.fill      = _fill(C_HEADER_BG)
        cell.alignment = _align("center", "center")
        cell.border    = _thin_border()
        _set_col_width(ws, col_idx, col_width)

    ws.freeze_panes = "A2"

    # ── Data rows ──
    for row_idx, result in enumerate(results, start=2):
        _set_row_height(ws, row_idx, 55)
        row_bg = C_ROW_ALT if row_idx % 2 == 0 else C_WHITE
        _write_result_row(ws, row_idx, row_idx - 1, result, row_bg)


def _write_result_row(ws, row: int, counter: int, result: dict, row_bg: str):
    api_type    = result.get("api_type", "")
    method      = result.get("method", "")
    status_code = result.get("status_code", 0)
    resp_time   = result.get("response_time_ms", 0)
    pass_fail   = result.get("pass_fail", "ERROR")

    # Col 1 — Row counter
    c = ws.cell(row=row, column=1, value=counter)
    c.font      = _font(bold=False, color="999999", size=9)
    c.fill      = _fill(C_GREY_BG)
    c.alignment = _align("center", "center")
    c.border    = _thin_border()

    # Col 2 — Test Case ID
    c = ws.cell(row=row, column=2, value=result.get("testcaseid", ""))
    c.font      = _font(bold=True, size=10, color=C_NAVY)
    c.fill      = _fill(row_bg)
    c.alignment = _align("left", "center")
    c.border    = _thin_border()

    # Col 3 — Type badge (REST/SOAP)
    type_bg = C_REST if api_type == "REST" else C_SOAP
    c = ws.cell(row=row, column=3, value=api_type)
    c.font      = _font(bold=True, size=9, color=C_WHITE)
    c.fill      = _fill(type_bg)
    c.alignment = _align("center", "center")
    c.border    = _thin_border()

    # Col 4 — Method badge
    method_bg = _method_color(method)
    c = ws.cell(row=row, column=4, value=method)
    c.font      = _font(bold=True, size=9, color=C_WHITE)
    c.fill      = _fill(method_bg)
    c.alignment = _align("center", "center")
    c.border    = _thin_border()

    # Col 5 — Endpoint
    c = ws.cell(row=row, column=5, value=result.get("endpoint", ""))
    c.font      = Font(name="Calibri", size=9, color="1565C0", underline="single")
    c.fill      = _fill(row_bg)
    c.alignment = _align("left", "center", wrap=True)
    c.border    = _thin_border()

    # Col 6 — Payload Sent (monospace)
    payload_text = str(result.get("payload_sent", "") or "")
    if len(payload_text) > 500:
        payload_text = payload_text[:500] + "…"
    c = ws.cell(row=row, column=6, value=payload_text)
    c.font      = Font(name="Courier New", size=8, color=C_NAVY)
    c.fill      = _fill(C_PAYLOAD_BG)
    c.alignment = _align("left", "top", wrap=True)
    c.border    = _thin_border()

    # Col 7 — Status Code
    code_bg = _status_color(status_code)
    code_val = status_code if status_code else "ERR"
    c = ws.cell(row=row, column=7, value=code_val)
    c.font      = _font(bold=True, size=11, color=C_NAVY)
    c.fill      = _fill(code_bg)
    c.alignment = _align("center", "center")
    c.border    = _thin_border()

    # Col 8 — Response Body (monospace)
    resp_text = str(result.get("response_body", "") or "")
    if len(resp_text) > 800:
        resp_text = resp_text[:800] + "…"
    c = ws.cell(row=row, column=8, value=resp_text)
    c.font      = Font(name="Courier New", size=8, color=C_NAVY)
    c.fill      = _fill(C_RESP_BG)
    c.alignment = _align("left", "top", wrap=True)
    c.border    = _thin_border()

    # Col 9 — Response Time
    time_bg = _time_color(resp_time)
    time_val = f"{resp_time} ms"
    c = ws.cell(row=row, column=9, value=time_val)
    c.font      = _font(bold=True, size=10, color=C_NAVY)
    c.fill      = _fill(time_bg)
    c.alignment = _align("center", "center")
    c.border    = _thin_border()

    # Col 10 — Status (PASS / FAIL / ERROR)
    if pass_fail == "PASS":
        pf_bg, pf_fg, pf_label = C_PASS_DARK, C_WHITE, "✅ PASS"
    elif pass_fail == "FAIL":
        pf_bg, pf_fg, pf_label = C_FAIL_DARK, C_WHITE, "❌ FAIL"
    else:
        pf_bg, pf_fg, pf_label = C_ERROR_DARK, C_WHITE, "⚠️ ERROR"
    c = ws.cell(row=row, column=10, value=pf_label)
    c.font      = Font(name="Calibri", bold=True, size=11, color=pf_fg)
    c.fill      = _fill(pf_bg)
    c.alignment = _align("center", "center")
    c.border    = _thin_border()


# ── Public entry point ────────────────────────────────────────────────────────

def generate_report(results: list, input_file: str, mode: str, output_path: str,
                    start_time: datetime = None, end_time: datetime = None) -> str:
    """
    Build and save the Excel report.
    """
    start_time = start_time or datetime.now()
    end_time   = end_time   or datetime.now()

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()

    # Sheet 1: Summary Dashboard
    ws_summary = wb.active
    ws_summary.title = "Summary Dashboard"
    _write_summary_sheet(ws_summary, results, input_file, mode, start_time, end_time)

    # Sheet 2: Test Results
    ws_results = wb.create_sheet("Test Results")
    _write_results_sheet(ws_results, results)

    # Print setup
    for ws in [ws_summary, ws_results]:
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToPage   = True
        ws.page_setup.fitToWidth  = 1

    wb.save(output_path)
    return output_path

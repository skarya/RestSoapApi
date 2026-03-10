"""
Script to generate Excel input files for the framework.
Run this once to create the .xlsx files from the JSON templates.
"""
import json
import sys
import os

# Add parent dir so we can import openpyxl after pip install
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR = os.path.join(BASE_DIR, "data", "input")

HEADER_COLOR = "1F2D3D"
HEADER_FONT  = "FFFFFF"
ROW_ALT      = "F2F6FC"

COLUMNS = [
    "TestCaseID", "Description", "Method", "Endpoint",
    "Headers", "Payload", "SOAPAction", "ExpectedStatus"
]


def style_header(cell):
    cell.font      = Font(name="Calibri", bold=True, color=HEADER_FONT, size=11)
    cell.fill      = PatternFill("solid", fgColor=HEADER_COLOR)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    cell.border    = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_data(cell, row_idx):
    bg = ROW_ALT if row_idx % 2 == 0 else "FFFFFF"
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    cell.border    = Border(left=thin, right=thin, top=thin, bottom=thin)


def json_to_excel(json_path: str, excel_path: str):
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32

    # Write header
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        style_header(cell)

    # Write data rows
    for row_idx, tc in enumerate(data, start=2):
        ws.row_dimensions[row_idx].height = 30
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            value = tc.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            style_data(cell, row_idx)

    # Auto column widths
    col_widths = [18, 35, 10, 45, 40, 35, 35, 15]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(excel_path)
    print(f"  ✅ Created: {os.path.basename(excel_path)}")


if __name__ == "__main__":
    print("\n  Generating Excel input files...\n")
    pairs = [
        ("TestSuite_REST.json",  "TestSuite_REST.xlsx"),
        ("TestSuite_SOAP.json",  "TestSuite_SOAP.xlsx"),
    ]
    for json_name, xlsx_name in pairs:
        json_path  = os.path.join(INPUT_DIR, json_name)
        excel_path = os.path.join(INPUT_DIR, xlsx_name)
        if os.path.isfile(json_path):
            json_to_excel(json_path, excel_path)
        else:
            print(f"  [SKIP] {json_name} not found.")
    print("\n  Done! Excel files in data/input/\n")
